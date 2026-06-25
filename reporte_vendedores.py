"""
Circa — Reporte diario de vendedores y bodegas
Uso:
  export SUPABASE_URL="https://rhxqcoijzgqlecpdfhde.supabase.co"
  export SUPABASE_SERVICE_KEY="eyJ..."
  python3 reporte_vendedores.py
"""

import os, sys
from datetime import datetime
import requests
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

SUPABASE_URL = os.environ.get("SUPABASE_URL", "").rstrip("/")
SUPABASE_KEY = os.environ.get("SUPABASE_SERVICE_KEY", "")
if not SUPABASE_URL or not SUPABASE_KEY:
    print("ERROR: Falta SUPABASE_URL y/o SUPABASE_SERVICE_KEY")
    sys.exit(1)

HEADERS = {"apikey": SUPABASE_KEY, "Authorization": f"Bearer {SUPABASE_KEY}"}


def sget(table, select="*", params=None):
    url = f"{SUPABASE_URL}/rest/v1/{table}"
    p = {"select": select}
    if params:
        p.update(params)
    r = requests.get(url, headers=HEADERS, params=p)
    if r.status_code != 200:
        print(f"ERROR {r.status_code}: {r.text}")
        sys.exit(1)
    return r.json()


def fetch_data():
    print("Descargando vendedores...")
    vendedores = sget("vendedores", "id,codigo,nombre,celular,activo", {"activo": "eq.true"})
    v_map = {v["id"]: v for v in vendedores}
    print(f"  {len(vendedores)} vendedores activos")

    print("Descargando bodega_vendedores...")
    bv_list = sget("bodega_vendedores", "bodega_id,vendedor_id,rol,grupo,supervisor,dia_visita,dia_entrega,activo", {"activo": "eq.true"})
    print(f"  {len(bv_list)} mapeos activos")

    bodega_ids = list(set(bv["bodega_id"] for bv in bv_list))

    print("Descargando bodegas...")
    bodegas_all = []
    for i in range(0, len(bodega_ids), 50):
        chunk = bodega_ids[i:i+50]
        b = sget(
            "bodegas",
            "id,razon_social,representante_nombre_corto,dni_representante,telefono_whatsapp,estado,kyc_nivel,onboarding_fase,linea_aprobada,linea_disponible,es_test,en_piloto",
            {"id": f"in.({','.join(chunk)})"}
        )
        bodegas_all.extend(b)
    b_map = {b["id"]: b for b in bodegas_all}
    print(f"  {len(bodegas_all)} bodegas")

    print("Descargando pedidos...")
    pedidos = sget("pedidos", "bodega_id,estado,created_at", {"estado": "in.(entregado,pagado,recibido,preventa_aceptada)"})
    ped_count, ped_last = {}, {}
    for p in pedidos:
        bid = p["bodega_id"]
        ped_count[bid] = ped_count.get(bid, 0) + 1
        ts = p.get("created_at", "")
        if ts > ped_last.get(bid, ""):
            ped_last[bid] = ts

    rows = []
    for bv in bv_list:
        v = v_map.get(bv["vendedor_id"])
        b = b_map.get(bv["bodega_id"])
        if not v or not b or b.get("es_test"):
            continue
        estado = b.get("estado", "?")
        last_raw = ped_last.get(b["id"], "")
        last_ped = ""
        if last_raw:
            try:
                last_ped = datetime.fromisoformat(last_raw.replace("Z", "+00:00")).strftime("%d/%m/%Y")
            except Exception:
                last_ped = last_raw[:10]
        rows.append((
            v.get("codigo", ""), v.get("nombre", ""), v.get("celular", ""),
            bv.get("supervisor", ""), bv.get("grupo", ""),
            bv.get("dia_visita", ""), bv.get("dia_entrega", ""),
            b.get("representante_nombre_corto") or b.get("razon_social") or "?",
            b.get("dni_representante", ""), b.get("telefono_whatsapp", ""),
            estado, "Sí" if estado == "activo" else "No",
            b.get("kyc_nivel") or "ninguno", b.get("onboarding_fase") or "invited",
            b.get("linea_aprobada") or 0, b.get("linea_disponible") or 0,
            ped_count.get(b["id"], 0), last_ped,
        ))
    rows.sort(key=lambda r: (r[3] or "", r[0] or "", 0 if r[11] == "Sí" else 1, r[7] or ""))
    return rows


HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=10)
HEADER_FILL = PatternFill("solid", fgColor="2B2B2B")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
ENROLADA_SI = PatternFill("solid", fgColor="D5F5E3")
ENROLADA_NO = PatternFill("solid", fgColor="FADBD8")
VENDOR_FILL = PatternFill("solid", fgColor="EBF5FB")
THIN_BORDER = Border(*(Side(style="thin", color="D5D8DC"),) * 4)
DATA_FONT = Font(name="Arial", size=9)

COLUMNS = [
    ("Código", 12), ("Vendedor", 25), ("Tel Vendedor", 14),
    ("Supervisor", 18), ("Grupo", 10), ("Día Visita", 12), ("Día Entrega", 12),
    ("Bodega", 30), ("DNI", 12), ("Tel Bodega", 15),
    ("Estado", 12), ("Enrolada", 10), ("KYC", 12), ("Onboarding", 16),
    ("Línea Aprobada", 15), ("Línea Disponible", 16),
    ("# Pedidos", 10), ("Último Pedido", 14),
]


def build_excel(rows):
    wb = Workbook()
    ws = wb.active
    ws.title = "Vendedores & Bodegas"

    for ci, (name, width) in enumerate(COLUMNS, 1):
        c = ws.cell(row=1, column=ci, value=name)
        c.font, c.fill, c.alignment, c.border = HEADER_FONT, HEADER_FILL, HEADER_ALIGN, THIN_BORDER
        ws.column_dimensions[get_column_letter(ci)].width = width
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}1"
    ws.freeze_panes = "A2"

    prev = None
    for ri, row in enumerate(rows, 2):
        new_v = row[0] != prev
        prev = row[0]
        for ci, val in enumerate(row, 1):
            c = ws.cell(row=ri, column=ci, value=val)
            c.font, c.border = DATA_FONT, THIN_BORDER
            if ci in (15, 16):
                c.number_format = '#,##0.00'
        if new_v:
            for ci in range(1, 8):
                ws.cell(row=ri, column=ci).fill = VENDOR_FILL
        ws.cell(row=ri, column=12).fill = ENROLADA_SI if row[11] == "Sí" else ENROLADA_NO

    # Resumen
    ws2 = wb.create_sheet("Resumen")
    sh = [("Código", 12), ("Vendedor", 25), ("Tel", 14), ("Supervisor", 18),
          ("Grupo", 10), ("Total", 10), ("Enroladas", 10), ("Pendientes", 10),
          ("Con Línea", 10), ("Con Pedidos", 10)]
    for ci, (name, width) in enumerate(sh, 1):
        c = ws2.cell(row=1, column=ci, value=name)
        c.font, c.fill, c.alignment, c.border = HEADER_FONT, HEADER_FILL, HEADER_ALIGN, THIN_BORDER
        ws2.column_dimensions[get_column_letter(ci)].width = width
    ws2.auto_filter.ref = f"A1:{get_column_letter(len(sh))}1"
    ws2.freeze_panes = "A2"

    vendors = {}
    for r in rows:
        k = r[0]
        if k not in vendors:
            vendors[k] = {"cod": r[0], "v": r[1], "t": r[2], "s": r[3], "g": r[4],
                          "tot": 0, "en": 0, "pe": 0, "cl": 0, "cp": 0}
        d = vendors[k]
        d["tot"] += 1
        if r[11] == "Sí": d["en"] += 1
        else: d["pe"] += 1
        if r[15] and r[15] > 0: d["cl"] += 1
        if r[16] and r[16] > 0: d["cp"] += 1

    for ri, d in enumerate(vendors.values(), 2):
        for ci, val in enumerate([d["cod"], d["v"], d["t"], d["s"], d["g"],
                                   d["tot"], d["en"], d["pe"], d["cl"], d["cp"]], 1):
            c = ws2.cell(row=ri, column=ci, value=val)
            c.font, c.border = DATA_FONT, THIN_BORDER

    fname = f"Circa_Reporte_Vendedores_{datetime.now().strftime('%Y%m%d')}.xlsx"
    wb.save(fname)
    return fname


def main():
    rows = fetch_data()
    print(f"  {len(rows)} filas")
    print(f"Reporte: {build_excel(rows)}")

if __name__ == "__main__":
    main()
