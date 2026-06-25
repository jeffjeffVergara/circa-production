#!/usr/bin/env python3
"""
CIRCA — Motor de Corte de Comisiones de Vendedores · v2.0
=========================================================
Un comando calcula las comisiones del periodo y genera un Excel claro
(agrupado por supervisor, vende arriba / no vende abajo) + mensajes de Mari.

CAMBIOS v2.0 (pedidos por Paola, jun-2026):
  1. Vendedores SIEMPRE con su supervisor (columna + agrupacion en el Excel).
  2. Comision EN VUELO: una comision se "materializa" cuando la bodega paga su
     pedido financiado. Hasta entonces va aparte, en AMARILLO. Verde = a pagar.
  3. Resumen ordenado: los que SI venden arriba, los que NO venden abajo.
  4. Total por vendedor con su pago final (solo lo MATERIALIZADO/verde).
  5. Top vendedor: estrella ⭐, NO se suma al pago (bono aparte que decide Paola).

USO:
    python3 circa_corte_comisiones.py --desde 2026-06-17 --hasta 2026-06-30

REQUISITOS (una vez):
    pip install requests openpyxl --break-system-packages
    export CIRCA_SUPABASE_URL="https://rhxqcoijzgqlecpdfhde.supabase.co"
    export CIRCA_SUPABASE_KEY="<service_role key>"

ESQUEMA (vigente 17-jun-2026):
    Afiliacion financiada S/5 | Afiliacion cash S/3 (1 vez por bodega)
    Recompra financiada S/2, max 4, primer mes de la bodega (cash NO paga)
    Meta semanal: 5 afiliaciones financiadas en la semana = S/10 (max 4/mes)
    Top: estrella, NO suma al pago.

ESTADOS DE COMISION:
    MATERIALIZADA (verde): bodega pago, o pedido cash -> a pagar.
    EN VUELO (amarillo): financiado, sin pagar, dentro de plazo -> aun no se paga.
    VENCIDA (rojo): financiado, sin pagar, vencido -> en espera hasta regularizar.

SEGURIDAD:
    Solo LECTURA. Excluye es_test=true y cuentas internas. Atribucion ambigua o
    sin vendedor -> hoja ALERTAS (nunca adivina).
"""

import argparse
import os
import sys
from collections import defaultdict
from datetime import date, datetime, timedelta

try:
    import requests
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
except ImportError as exc:
    sys.exit(f"FALTA UNA LIBRERIA: {exc}. Corre: pip install requests openpyxl --break-system-packages")

# ----------------------------- Configuracion ------------------------------
TARIFAS = {"afiliacion_financiada": 5.0, "afiliacion_cash": 3.0,
           "recompra_financiada": 2.0, "meta_semanal": 10.0}
MAX_RECOMPRAS_MES1 = 4
META_SEMANAL_AFILIACIONES = 5
MAX_METAS_POR_PERIODO = 4
DIAS_PRIMER_MES = 30
DIMAX_DISTRIBUIDOR_ID = "d1a2b3c4-0001-4000-8000-000000000002"
CODIGOS_INTERNOS = {"VW-PAO", "VW-CYNTHIA", "CIRCA01", "VWTEST-JV", "VW-SARASVATI"}
ESTADOS_EXCLUIDOS_FRAGMENTOS = ("cancel", "rechaz", "expir", "anulad")

VERDE, VERDE_TXT = "C6EFCE", "0E6B3A"
AMARILLO, AMARILLO_TXT = "FFEB9C", "9C6500"
ROJO, ROJO_TXT = "FFC7CE", "9C0006"
AZUL, GRIS = "2F56C9", "E7ECF5"

# ------------------------------- Supabase ---------------------------------
def env_o_falla(n):
    v = os.environ.get(n, "").strip()
    if not v:
        sys.exit(f"ERROR: falta la variable de entorno {n}. Ver cabecera del script.")
    return v


def fetch_todo(base_url, key, tabla, select, filtros=""):
    headers = {"apikey": key, "Authorization": f"Bearer {key}", "Prefer": "count=exact"}
    filas, offset, page = [], 0, 1000
    while True:
        url = f"{base_url}/rest/v1/{tabla}?select={select}"
        if filtros:
            url += f"&{filtros}"
        headers["Range"] = f"{offset}-{offset + page - 1}"
        r = requests.get(url, headers=headers, timeout=60)
        if r.status_code not in (200, 206):
            sys.exit(f"ERROR Supabase '{tabla}' (HTTP {r.status_code}): {r.text[:400]}")
        lote = r.json()
        filas.extend(lote)
        if len(lote) < page:
            return filas
        offset += page


def parse_ts(v):
    if not v:
        return None
    try:
        return datetime.fromisoformat(str(v).replace("Z", "+00:00"))
    except ValueError:
        return None


def parse_d(v):
    ts = parse_ts(v)
    if ts:
        return ts.date()
    try:
        return date.fromisoformat(str(v)[:10])
    except (ValueError, TypeError):
        return None


def nombre_pila(v):
    # Prefiere el nombre_corto explicito de la base (siempre correcto).
    # Si no existe, cae a una regla automatica (APELLIDO APELLIDO NOMBRES).
    corto = (v.get("nombre_corto") or "").strip()
    if corto:
        return corto
    t = (v.get("nombre") or "").split()
    if len(t) >= 3:
        return f"{t[2]} {t[0]}".title()
    if len(t) == 2:
        return f"{t[1]} {t[0]}".title()
    return (t[0] if t else v.get("codigo", "")).title()


# ------------------------------- Main -------------------------------------
def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--desde", required=True)
    ap.add_argument("--hasta", required=True)
    args = ap.parse_args()
    desde, hasta = date.fromisoformat(args.desde), date.fromisoformat(args.hasta)
    if hasta < desde:
        sys.exit("ERROR: --hasta es anterior a --desde.")
    hoy = date.today()

    base_url = env_o_falla("CIRCA_SUPABASE_URL").rstrip("/")
    key = env_o_falla("CIRCA_SUPABASE_KEY")
    print(f"Corte de comisiones {desde} a {hasta} (solo lectura)\n")

    vendedores = fetch_todo(base_url, key, "vendedores",
                            "id,codigo,nombre,celular,activo,supervisor,nombre_corto",
                            f"distribuidor_id=eq.{DIMAX_DISTRIBUIDOR_ID}&activo=eq.true")
    vendedores = [v for v in vendedores if v["codigo"] not in CODIGOS_INTERNOS]
    if not vendedores:
        sys.exit("ERROR: 0 vendedores reales activos.")
    vend = {v["id"]: v for v in vendedores}
    sin_sup = [v["codigo"] for v in vendedores if not (v.get("supervisor") or "").strip()]
    print(f"  Vendedores reales activos: {len(vendedores)}")
    if sin_sup:
        print(f"  AVISO sin supervisor: {', '.join(sin_sup)}")

    bv = fetch_todo(base_url, key, "bodega_vendedores", "bodega_id,vendedor_id,activo,rol", "activo=eq.true")
    por_bod = defaultdict(list)
    for f in bv:
        if f["vendedor_id"] in vend:
            por_bod[f["bodega_id"]].append(f)
    bod_a_vend, ambiguas = {}, set()
    for bid, mps in por_bod.items():
        if len(mps) == 1:
            bod_a_vend[bid] = mps[0]["vendedor_id"]
        else:
            no_conf = [m for m in mps if (m.get("rol") or "").upper() != "CONFITERIA"]
            if len(no_conf) == 1:
                bod_a_vend[bid] = no_conf[0]["vendedor_id"]
            else:
                ambiguas.add(bid)
    print(f"  Mapeos a vendedores reales: {len(bv)} ({len(ambiguas)} ambiguas)")

    bodegas = fetch_todo(base_url, key, "bodegas", "id,nombre_comercial,razon_social,es_test", "es_test=eq.false")
    bod = {b["id"]: b for b in bodegas}
    print(f"  Bodegas reales: {len(bodegas)}")

    pedidos = fetch_todo(base_url, key, "pedidos",
                         "id,numero,bodega_id,vendedor_id,monto_financiado,monto_contado,total,"
                         "estado,origen,created_at,pagado_at,fecha_pagado,fecha_vencimiento",
                         f"distribuidor_id=eq.{DIMAX_DISTRIBUIDOR_ID}")
    pedidos = [p for p in pedidos if p["bodega_id"] in bod]
    estados = sorted({(p.get("estado") or "NULL") for p in pedidos})
    excl = {e for e in estados if any(f in e.lower() for f in ESTADOS_EXCLUIDOS_FRAGMENTOS)}
    print(f"\n  Estados: {', '.join(estados)}")
    if excl:
        print(f"  Excluidos: {', '.join(sorted(excl))}")
    pedidos = [p for p in pedidos if (p.get("estado") or "NULL") not in excl]
    print(f"  Pedidos validos: {len(pedidos)}")

    for p in pedidos:
        p["_f"] = parse_ts(p["created_at"]).date() if parse_ts(p["created_at"]) else None
        p["_fin"] = float(p["monto_financiado"] or 0) > 0
        p["_pag"] = parse_ts(p.get("pagado_at")) or parse_ts(p.get("fecha_pagado"))
        p["_venc"] = parse_d(p.get("fecha_vencimiento"))

    pb = defaultdict(list)
    for p in pedidos:
        if p["_f"]:
            pb[p["bodega_id"]].append(p)
    for l in pb.values():
        l.sort(key=lambda x: x["created_at"])

    def estado_com(p):
        if not p["_fin"]:
            return "MAT"
        if p["_pag"] is not None:
            return "MAT"
        if p["_venc"] and p["_venc"] < hoy:
            return "VENC"
        return "VUELO"

    def atribuir(p):
        return p.get("vendedor_id") if p.get("vendedor_id") in vend else bod_a_vend.get(p["bodega_id"])

    detalle, alertas = [], []
    paga = defaultdict(lambda: defaultdict(float))
    cnt = defaultdict(lambda: defaultdict(int))
    afil_sem = defaultdict(lambda: defaultdict(int))
    afil_periodo = defaultdict(int)

    def reg(vid, nb, p, concepto, monto):
        est = estado_com(p)
        paga[vid][est] += monto
        cnt[vid][concepto] += 1
        detalle.append((vend[vid].get("supervisor") or "SIN SUPERVISOR", vend[vid]["codigo"],
                        nb, p.get("numero") or "", p["_f"], concepto.replace("_", " "), monto, est))

    for bid, lst in pb.items():
        fin_mes1 = lst[0]["_f"] + timedelta(days=DIAS_PRIMER_MES)
        nb = bod[bid].get("nombre_comercial") or bod[bid].get("razon_social") or bid[:8]
        recompras = 0
        for idx, p in enumerate(lst):
            vid = atribuir(p)
            if vid is None:
                if desde <= p["_f"] <= hasta:
                    motivo = "ATRIBUCION AMBIGUA" if bid in ambiguas else "SIN VENDEDOR asignado"
                    alertas.append((p.get("numero") or p["id"][:8], nb, p["_f"], motivo))
                continue
            en = desde <= p["_f"] <= hasta
            if idx == 0:
                c = "afiliacion_financiada" if p["_fin"] else "afiliacion_cash"
                if en:
                    afil_periodo[vid] += 1
                    if p["_fin"]:
                        afil_sem[vid][p["_f"].isocalendar()[:2]] += 1
                    reg(vid, nb, p, c, TARIFAS[c])
            else:
                if not p["_fin"]:
                    continue
                if p["_f"] <= fin_mes1:
                    if recompras >= MAX_RECOMPRAS_MES1:
                        if en:
                            detalle.append((vend[vid].get("supervisor") or "SIN SUPERVISOR", vend[vid]["codigo"],
                                            nb, p.get("numero") or "", p["_f"], "recompra (sobre tope 4)", 0.0, "NOPAGA"))
                        continue
                    recompras += 1
                    if en:
                        reg(vid, nb, p, "recompra_financiada", TARIFAS["recompra_financiada"])
                elif en:
                    cnt[vid]["bolsa_mes2"] += 1
                    detalle.append((vend[vid].get("supervisor") or "SIN SUPERVISOR", vend[vid]["codigo"],
                                    nb, p.get("numero") or "", p["_f"], "recompra mes 2+ (bolsa)", 0.0, "BOLSA"))

    for vid, sem in afil_sem.items():
        metas = min(sum(1 for c in sem.values() if c >= META_SEMANAL_AFILIACIONES), MAX_METAS_POR_PERIODO)
        if metas:
            cnt[vid]["metas_semanales"] = metas
            paga[vid]["MAT"] += metas * TARIFAS["meta_semanal"]

    top_vid = None
    if afil_periodo:
        top_n = max(afil_periodo.values())
        tops = [v for v, n in afil_periodo.items() if n == top_n and n > 0]
        top_vid = tops[0] if len(tops) == 1 else None
        if len(tops) > 1:
            print(f"\n  AVISO empate top ({len(tops)} con {top_n} afil.). Decidir manual.")

    # ------------------------------- Excel --------------------------------
    base = f"Corte_Comisiones_{desde}_{hasta}"
    wb = Workbook()
    thin = Side(style="thin", color="D0D7E5")
    borde = Border(left=thin, right=thin, top=thin, bottom=thin)
    hf = PatternFill("solid", fgColor=AZUL)
    ht = Font(color="FFFFFF", bold=True, size=11)
    sf = PatternFill("solid", fgColor=GRIS)

    ws = wb.active
    ws.title = "Resumen"
    ws.merge_cells("A1:J1")
    ws["A1"] = f"CORTE DE COMISIONES  ·  {desde.strftime('%d/%m/%Y')} a {hasta.strftime('%d/%m/%Y')}"
    ws["A1"].font = Font(bold=True, size=14, color=AZUL)
    ws.row_dimensions[1].height = 26
    ws.merge_cells("A2:J2")
    ws["A2"] = ("Verde = a pagar (bodega ya pago)   |   Amarillo = en vuelo (se paga si la bodega paga)"
                "   |   Rojo = en espera (no pago a tiempo)   |   ⭐ = top (bono aparte)")
    ws["A2"].font = Font(italic=True, size=9, color="5B6678")

    cab = ["Vendedor", "Cod.", "Afil.fin", "Afil.cash", "Recompras", "Metas",
           "A PAGAR S/", "EN VUELO S/", "VENCIDO S/", "⭐"]

    def cabecera(f):
        for j, h in enumerate(cab, 1):
            c = ws.cell(row=f, column=j, value=h)
            c.fill, c.font, c.border = hf, ht, borde
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.row_dimensions[f].height = 30

    def fila_vend(f, v):
        vid = v["id"]
        mat, vu, ve = round(paga[vid]["MAT"], 2), round(paga[vid]["VUELO"], 2), round(paga[vid]["VENC"], 2)
        vals = [nombre_pila(v), v["codigo"], cnt[vid]["afiliacion_financiada"],
                cnt[vid]["afiliacion_cash"], cnt[vid]["recompra_financiada"], cnt[vid]["metas_semanales"],
                mat, vu, ve, "⭐" if vid == top_vid else ""]
        for j, val in enumerate(vals, 1):
            c = ws.cell(row=f, column=j, value=val)
            c.border = borde
            c.alignment = Alignment(horizontal="center" if j > 1 else "left", vertical="center")
        ws.cell(row=f, column=7).fill = PatternFill("solid", fgColor=VERDE)
        ws.cell(row=f, column=7).font = Font(bold=True, color=VERDE_TXT)
        if vu > 0:
            ws.cell(row=f, column=8).fill = PatternFill("solid", fgColor=AMARILLO)
            ws.cell(row=f, column=8).font = Font(color=AMARILLO_TXT)
        if ve > 0:
            ws.cell(row=f, column=9).fill = PatternFill("solid", fgColor=ROJO)
            ws.cell(row=f, column=9).font = Font(color=ROJO_TXT)
        return mat, vu, ve

    def total_v(v):
        vid = v["id"]
        return (paga[vid]["MAT"] + paga[vid]["VUELO"] + paga[vid]["VENC"]
                + cnt[vid]["afiliacion_financiada"] + cnt[vid]["afiliacion_cash"] + cnt[vid]["recompra_financiada"])

    venden = sorted([v for v in vendedores if total_v(v) > 0],
                    key=lambda v: (v.get("supervisor") or "ZZZ", -paga[v["id"]]["MAT"]))
    no_venden = sorted([v for v in vendedores if total_v(v) == 0],
                       key=lambda v: (v.get("supervisor") or "ZZZ", v["codigo"]))

    f = 4
    tm = tv = tx = 0.0
    ws.cell(row=f, column=1, value="✅ VENDEDORES CON ACTIVIDAD").font = Font(bold=True, size=12, color=VERDE_TXT)
    f += 1
    cabecera(f); f += 1
    sup = None
    for v in venden:
        s = v.get("supervisor") or "SIN SUPERVISOR"
        if s != sup:
            ws.merge_cells(start_row=f, start_column=1, end_row=f, end_column=10)
            c = ws.cell(row=f, column=1, value=f"  Supervisor: {s}")
            c.fill, c.font = sf, Font(bold=True, color=AZUL)
            f += 1; sup = s
        m, vu, ve = fila_vend(f, v)
        tm += m; tv += vu; tx += ve; f += 1

    f += 1
    ws.cell(row=f, column=1, value="⬜ SIN ACTIVIDAD EN EL PERIODO").font = Font(bold=True, size=12, color="8A93A6")
    f += 1
    cabecera(f); f += 1
    sup = None
    for v in no_venden:
        s = v.get("supervisor") or "SIN SUPERVISOR"
        if s != sup:
            ws.merge_cells(start_row=f, start_column=1, end_row=f, end_column=10)
            c = ws.cell(row=f, column=1, value=f"  Supervisor: {s}")
            c.fill, c.font = sf, Font(bold=True, color="8A93A6")
            f += 1; sup = s
        fila_vend(f, v); f += 1

    f += 1
    ws.cell(row=f, column=1, value="TOTAL EQUIPO").font = Font(bold=True, size=12)
    for col, val, color in [(7, tm, VERDE), (8, tv, AMARILLO), (9, tx, ROJO)]:
        c = ws.cell(row=f, column=col, value=round(val, 2))
        c.fill = PatternFill("solid", fgColor=color)
        c.font = Font(bold=True, size=12)
        c.border = borde

    for j, a in enumerate([26, 8, 9, 9, 11, 8, 13, 13, 12, 5], 1):
        ws.column_dimensions[get_column_letter(j)].width = a
    ws.freeze_panes = "A4"

    # Detalle
    wd = wb.create_sheet("Detalle")
    for j, h in enumerate(["Supervisor", "Vendedor", "Bodega", "Pedido", "Fecha", "Concepto", "S/", "Estado"], 1):
        c = wd.cell(row=1, column=j, value=h); c.fill, c.font = hf, ht
    rr = 2
    estado_legible = {"MAT": "A PAGAR", "VUELO": "EN VUELO", "VENC": "VENCIDO", "BOLSA": "BOLSA M2+", "NOPAGA": "SOBRE TOPE"}
    for d in sorted(detalle):
        s, cod, b, p, fch, con, mon, est = d
        for j, val in enumerate([s, cod, b, p, fch.isoformat() if fch else "", con, mon, estado_legible.get(est, est)], 1):
            wd.cell(row=rr, column=j, value=val)
        ce = wd.cell(row=rr, column=8)
        if est == "MAT":
            ce.fill = PatternFill("solid", fgColor=VERDE); ce.font = Font(color=VERDE_TXT)
        elif est == "VUELO":
            ce.fill = PatternFill("solid", fgColor=AMARILLO); ce.font = Font(color=AMARILLO_TXT)
        elif est == "VENC":
            ce.fill = PatternFill("solid", fgColor=ROJO); ce.font = Font(color=ROJO_TXT)
        rr += 1
    for j, a in enumerate([26, 10, 28, 14, 11, 26, 7, 12], 1):
        wd.column_dimensions[get_column_letter(j)].width = a
    wd.freeze_panes = "A2"

    # Alertas
    wa = wb.create_sheet("Alertas")
    for j, h in enumerate(["Pedido", "Bodega", "Fecha", "Problema"], 1):
        c = wa.cell(row=1, column=j, value=h); c.fill, c.font = hf, ht
    for i, al in enumerate(alertas, 2):
        wa.cell(row=i, column=1, value=al[0]); wa.cell(row=i, column=2, value=al[1])
        wa.cell(row=i, column=3, value=al[2].isoformat() if al[2] else ""); wa.cell(row=i, column=4, value=al[3])
    for j, a in enumerate([16, 28, 11, 40], 1):
        wa.column_dimensions[get_column_letter(j)].width = a

    wb.save(f"{base}.xlsx")

    # Mensajes de Mari
    lineas = []
    for v in sorted(vendedores, key=lambda x: (x.get("supervisor") or "ZZZ", x["codigo"])):
        vid = v["id"]
        mat, vu, ve = round(paga[vid]["MAT"], 2), round(paga[vid]["VUELO"], 2), round(paga[vid]["VENC"], 2)
        n = nombre_pila(v)
        if mat == 0 and vu == 0 and ve == 0 and cnt[vid]["bolsa_mes2"] == 0:
            msg = (f"Hola {n} 👋 Soy Mari. Esta semana aun no registro bodegas tuyas en Circa. "
                   f"Dime cual de tu ruta quieres activar primero y te preparo todo: linea, promos y que decirle.")
        else:
            partes = []
            if cnt[vid]["afiliacion_financiada"]:
                partes.append(f"{cnt[vid]['afiliacion_financiada']} afiliacion(es) financiada(s)")
            if cnt[vid]["afiliacion_cash"]:
                partes.append(f"{cnt[vid]['afiliacion_cash']} afiliacion(es) cash")
            if cnt[vid]["recompra_financiada"]:
                partes.append(f"{cnt[vid]['recompra_financiada']} recompra(s) financiada(s)")
            if cnt[vid]["metas_semanales"]:
                partes.append(f"{cnt[vid]['metas_semanales']} meta(s) semanal(es)")
            msg = (f"Hola {n} 👋 Soy Mari. Tu avance Circa del {desde.strftime('%d/%m')} al {hasta.strftime('%d/%m')}: "
                   + " + ".join(partes) + f". Llevas S/{mat:.2f} confirmados a pagar.")
            if vu:
                msg += f" Y S/{vu:.2f} en camino — se confirman cuando esas bodegas paguen su pedido."
            if ve:
                msg += f" Ojo: S/{ve:.2f} en espera por bodegas que no pagaron a tiempo — ayudalas a ponerse al dia."
            if vid == top_vid:
                msg += " 🏆 ¡Vas TOP del periodo!"
            msg += " El pago sale a fin de mes por Yape. 💪"
        lineas.append(f"=== {v['codigo']} · {v['nombre']} · {v.get('celular') or 'SIN CELULAR'} "
                      f"· Sup: {v.get('supervisor') or '-'} ===\n{msg}\n")
    with open(f"Mensajes_{base}.txt", "w", encoding="utf-8") as fh:
        fh.write("\n".join(lineas))

    print(f"\nLISTO:")
    print(f"  {base}.xlsx  (Resumen + Detalle + {len(alertas)} alertas)")
    print(f"  Mensajes_{base}.txt  ({len(vendedores)} mensajes)")
    print(f"\n  A PAGAR (confirmado): S/{tm:.2f}")
    print(f"  EN VUELO (si pagan):  S/{tv:.2f}")
    print(f"  VENCIDO (en espera):  S/{tx:.2f}")
    if top_vid:
        print(f"  ⭐ Top: {vend[top_vid]['codigo']} (bono aparte)")
    if alertas:
        print(f"\n  ATENCION: {len(alertas)} alertas — resolver ANTES de pagar.")


if __name__ == "__main__":
    main()
