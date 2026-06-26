"""
Circa — Import DIMAX bodega Excel con modelo de riesgo.
Endpoint: POST /api/backoffice/import/dimax-analisis/preview
          POST /api/backoffice/import/dimax-analisis/confirm

Flujo:
1. Upload Excel DIMAX (hoja clientes + Historial)
2. Backend parsea, corre modelo de riesgo, resuelve vendedor
3. Devuelve preview con campos editables
4. Frontend confirma → crea bodega + vendedor + mapping
"""
from __future__ import annotations

import io
import logging
import re
from collections import defaultdict
from datetime import datetime, timedelta
from typing import Any, Optional

from fastapi import Depends, File, HTTPException, UploadFile
from pydantic import BaseModel

from app.services import db
from app.services.backoffice_auth import get_backoffice_user, get_backoffice_writer, verify_reauth_password

logger = logging.getLogger("circa.import_dimax")

DIMAX_ID = "d1a2b3c4-0001-4000-8000-000000000002"


# ── Helpers ──────────────────────────────────────────────

def _telefono_e164(raw):
    if not raw:
        return None
    s = re.sub(r"[^\d]", "", str(raw))
    if len(s) == 9 and s[0] == "9":
        return f"+51{s}"
    if len(s) == 11 and s[:2] == "51":
        return f"+{s}"
    if len(s) == 12 and s[:3] == "051":
        return f"+51{s[3:]}"
    return f"+51{s}" if len(s) == 9 else None


def _titulo(s):
    if not s:
        return ""
    return " ".join(w.capitalize() if w.islower() else w for w in str(s).strip().split())


def _doc_info(raw):
    s = re.sub(r"[^\d]", "", str(raw or ""))
    if len(s) == 11:
        return {"ruc": s, "dni": None, "solo_dni": False}
    elif len(s) >= 7:
        return {"ruc": None, "dni": s, "solo_dni": True}
    return {"ruc": None, "dni": s or None, "solo_dni": True}


def _dia_min(raw):
    if not raw:
        return ""
    d = str(raw).strip().lower()
    mapa = {"lunes": "lunes", "martes": "martes", "miercoles": "miercoles",
            "miércoles": "miercoles", "jueves": "jueves", "viernes": "viernes",
            "sabado": "sabado", "sábado": "sabado", "domingo": "domingo"}
    return mapa.get(d, d)


def _parse_fecha(raw):
    if not raw:
        return None
    if isinstance(raw, datetime):
        return raw
    try:
        return datetime.strptime(str(raw)[:10], "%Y-%m-%d")
    except (ValueError, TypeError):
        return None


# ── Modelo de riesgo ─────────────────────────────────────

def analizar_riesgo(historial_rows: list[dict], hoy: datetime = None) -> dict:
    """Mismo modelo que cargar_bodega.py"""
    hoy = hoy or datetime(2026, 6, 26)
    hace_6m = hoy - timedelta(days=180)

    pedidos_por_fecha = defaultdict(float)
    for row in historial_rows:
        fecha = _parse_fecha(row.get("FechaFacturacion") or row.get("fecha"))
        monto = row.get("SELL OUT") or row.get("monto") or 0
        if not fecha:
            continue
        try:
            m = float(monto)
        except (ValueError, TypeError):
            continue
        if fecha >= hace_6m:
            pedidos_por_fecha[fecha.strftime("%Y-%m-%d")] += m

    fechas = sorted(pedidos_por_fecha.keys())
    n_pedidos = len(fechas)
    total = sum(pedidos_por_fecha.values())
    ticket = total / n_pedidos if n_pedidos > 0 else 0

    gaps = []
    for i in range(1, len(fechas)):
        d1 = datetime.strptime(fechas[i - 1], "%Y-%m-%d")
        d2 = datetime.strptime(fechas[i], "%Y-%m-%d")
        gaps.append((d2 - d1).days)

    avg_gap = sum(gaps) / len(gaps) if gaps else 0
    import statistics
    cv = (statistics.stdev(gaps) / avg_gap) if gaps and avg_gap > 0 and len(gaps) > 1 else 0
    regularidad = ("muy regular" if cv < 0.3 else "regular" if cv < 0.6
                   else "irregular" if cv < 1.0 else "muy irregular")

    fecha_min = fechas[0] if fechas else ""
    fecha_max = fechas[-1] if fechas else ""
    dias_periodo = 1
    if fecha_min and fecha_max:
        dias_periodo = max(1, (datetime.strptime(fecha_max, "%Y-%m-%d") -
                               datetime.strptime(fecha_min, "%Y-%m-%d")).days)
    consumo_diario = total / dias_periodo
    linea_7d = consumo_diario * 7

    if linea_7d <= 100:
        tier = 100
    elif linea_7d <= 200:
        tier = 200
    elif linea_7d <= 300:
        tier = 300
    elif linea_7d <= 400:
        tier = 400
    else:
        tier = 500

    return {
        "periodo_desde": fecha_min,
        "periodo_hasta": fecha_max,
        "n_pedidos": n_pedidos,
        "total_comprado": round(total, 2),
        "ticket_promedio": round(ticket, 2),
        "dias_entre_pedidos": round(avg_gap, 1),
        "regularidad": regularidad,
        "cv": round(cv, 2),
        "consumo_diario": round(consumo_diario, 2),
        "linea_7d": round(linea_7d, 2),
        "tier_sugerido": tier,
    }


# ── Resolver vendedor ────────────────────────────────────

def _resolver_vendedor_circa(cod_dimax: str) -> dict | None:
    """Busca vendedor en tabla vendedores por código DIMAX."""
    if not cod_dimax:
        return None
    rows = (db.sb.table("vendedores")
            .select("id, codigo, nombre, activo")
            .eq("codigo", cod_dimax.strip())
            .limit(1).execute().data or [])
    if rows:
        return rows[0]
    # Si no existe con ese código exacto, crear automáticamente al confirmar
    return None


# ── Parse Excel DIMAX ────────────────────────────────────

def _parse_dimax_excel(content: bytes) -> dict:
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)

    # Buscar hoja clientes
    clientes_sheet = None
    historial_sheet = None
    for name in wb.sheetnames:
        nl = name.lower().strip()
        if "cliente" in nl:
            clientes_sheet = wb[name]
        elif "historial" in nl:
            historial_sheet = wb[name]

    if not clientes_sheet:
        raise HTTPException(400, "No se encontró hoja 'clientes' en el Excel")

    # Parse headers
    headers = [str(clientes_sheet.cell(1, c).value or "").strip()
               for c in range(1, clientes_sheet.max_column + 1)]
    cliente = {}
    for c, h in enumerate(headers, 1):
        cliente[h] = clientes_sheet.cell(2, c).value

    # Parse historial
    hist_rows = []
    if historial_sheet and historial_sheet.max_row > 1:
        h_headers = [str(historial_sheet.cell(1, c).value or "").strip()
                     for c in range(1, historial_sheet.max_column + 1)]
        for r in range(2, historial_sheet.max_row + 1):
            row = {}
            for c, h in enumerate(h_headers, 1):
                row[h] = historial_sheet.cell(r, c).value
            hist_rows.append(row)

    return {"cliente": cliente, "historial": hist_rows}


# ── Preview endpoint ─────────────────────────────────────

async def import_dimax_preview_handler(
    file: UploadFile = File(...),
    user: dict = Depends(get_backoffice_user),
):
    content = await file.read()
    if len(content) > 10 * 1024 * 1024:
        raise HTTPException(400, "Archivo muy grande (máx 10MB)")

    parsed = _parse_dimax_excel(content)
    c = parsed["cliente"]
    hist = parsed["historial"]

    # Info del cliente
    tel = _telefono_e164(c.get("TELEFONO"))
    doc = _doc_info(c.get("DNI/RUC"))
    razon = str(c.get("RazonSocial", "")).strip()
    direccion = _titulo(c.get("Direccion"))
    distrito = _titulo(c.get("Distrito"))

    # Vendedor - buscar en columnas con y sin sufijo
    cod_vendedor = (c.get("COD VENDEDOR") or c.get("COD VENDEDOR 1") or "")
    nom_vendedor = (c.get("VENDEDOR") or c.get("VENDEDOR 1") or "")
    grupo = (c.get("GRUPO") or c.get("GRUPO 1") or "")
    supervisor = (c.get("SUPERVISOR") or c.get("SUPERVISOR 1") or "")
    dia_visita = _dia_min(c.get("DIA VISITA") or c.get("DIA VISITA 1"))
    dia_entrega = _dia_min(c.get("DIA ENTREGA") or c.get("DIA ENTREGA 1"))

    # Resolver vendedor en BD
    v_circa = _resolver_vendedor_circa(str(cod_vendedor).strip()) if cod_vendedor else None

    # Verificar si bodega ya existe
    existe = False
    if tel:
        existing = (db.sb.table("bodegas")
                    .select("id, nombre_comercial, estado, linea_aprobada")
                    .eq("telefono_whatsapp", tel)
                    .limit(1).execute().data)
        if existing:
            existe = True

    # Correr modelo de riesgo
    riesgo = analizar_riesgo(hist)

    return {
        "ok": True,
        "filename": file.filename,
        "bodega_existe": existe,
        "cliente": {
            "razon_social": razon,
            "nombre_comercial": razon,
            "telefono_whatsapp": tel,
            "ruc": doc["ruc"],
            "dni_representante": doc["dni"],
            "solo_dni_sin_ruc": doc["solo_dni"],
            "direccion_fiscal": direccion,
            "distrito": distrito,
            "codigo_dimax": str(c.get("Codigo", "")).strip(),
            "clasificacion": str(c.get("Clasificacion", "")).strip(),
        },
        "vendedor": {
            "cod_dimax": str(cod_vendedor).strip() if cod_vendedor else "",
            "nombre_dimax": str(nom_vendedor).strip() if nom_vendedor else "",
            "grupo": str(grupo).strip() if grupo else "",
            "supervisor": str(supervisor).strip() if supervisor else "",
            "dia_visita": dia_visita,
            "dia_entrega": dia_entrega,
            "vendedor_circa": v_circa,  # None si no existe, {id, codigo, nombre} si existe
        },
        "riesgo": riesgo,
        "linea_sugerida": riesgo["tier_sugerido"],
    }


# ── Confirm endpoint ─────────────────────────────────────

class DiMaxConfirmBody(BaseModel):
    razon_social: str
    nombre_comercial: Optional[str] = None
    telefono_whatsapp: str
    ruc: Optional[str] = None
    dni_representante: Optional[str] = None
    solo_dni_sin_ruc: bool = True
    direccion_fiscal: Optional[str] = None
    distrito: Optional[str] = None
    linea_aprobada: float = 100
    # Vendedor
    cod_vendedor: Optional[str] = None
    nombre_vendedor: Optional[str] = None
    grupo: Optional[str] = None
    supervisor: Optional[str] = None
    dia_visita: Optional[str] = None
    dia_entrega: Optional[str] = None
    # Auth
    password: str
    comentario: str = "Carga DIMAX con análisis"


async def import_dimax_confirm_handler(
    body: DiMaxConfirmBody,
    user: dict = Depends(get_backoffice_writer),
):
    verify_reauth_password(body.password)

    tel = body.telefono_whatsapp
    if not tel:
        raise HTTPException(400, "Teléfono requerido")

    # Verificar si ya existe
    existing = (db.sb.table("bodegas")
                .select("id")
                .eq("telefono_whatsapp", tel)
                .limit(1).execute().data)

    bodega_id = None
    if existing:
        bodega_id = existing[0]["id"]
        # Actualizar línea si cambió
        db.sb.table("bodegas").update({
            "linea_aprobada": body.linea_aprobada,
        }).eq("id", bodega_id).execute()
        logger.info(f"Bodega existente actualizada: {bodega_id}")
    else:
        # Crear bodega
        insert_data = {
            "distribuidor_id": DIMAX_ID,
            "razon_social": body.razon_social,
            "nombre_comercial": body.nombre_comercial or body.razon_social,
            "telefono_whatsapp": tel,
            "ruc": body.ruc,
            "dni_representante": body.dni_representante,
            "solo_dni_sin_ruc": body.solo_dni_sin_ruc,
            "direccion_fiscal": body.direccion_fiscal,
            "distrito": body.distrito,
            "es_test": False,
            "en_piloto": True,
            "estado": "inactivo",
            "linea_aprobada": body.linea_aprobada,
            "linea_disponible": 0,
        }
        result = db.sb.table("bodegas").insert(insert_data).execute()
        if result.data:
            bodega_id = result.data[0]["id"]
        logger.info(f"Bodega creada: {bodega_id}")

    if not bodega_id:
        raise HTTPException(500, "Error al crear/encontrar bodega")

    # Crear vendedor si no existe
    vendedor_id = None
    if body.cod_vendedor:
        v_rows = (db.sb.table("vendedores")
                  .select("id")
                  .eq("codigo", body.cod_vendedor)
                  .limit(1).execute().data)
        if v_rows:
            vendedor_id = v_rows[0]["id"]
        else:
            # Crear vendedor
            v_result = db.sb.table("vendedores").insert({
                "distribuidor_id": DIMAX_ID,
                "codigo": body.cod_vendedor,
                "nombre": body.nombre_vendedor or body.cod_vendedor,
                "activo": True,
            }).execute()
            if v_result.data:
                vendedor_id = v_result.data[0]["id"]
            logger.info(f"Vendedor creado: {body.cod_vendedor} -> {vendedor_id}")

    # Crear mapping bodega_vendedores
    if vendedor_id:
        # Verificar si ya existe
        existing_bv = (db.sb.table("bodega_vendedores")
                       .select("id")
                       .eq("bodega_id", bodega_id)
                       .eq("vendedor_id", vendedor_id)
                       .limit(1).execute().data)
        if not existing_bv:
            grupo_norm = body.grupo or ""
            if grupo_norm.startswith("GV "):
                rol = grupo_norm.replace("GV ", "").split()[0] if " " in grupo_norm else "ABN"
            elif "MERCADO" in grupo_norm.upper():
                rol = "MERCADOS"
            else:
                rol = "ABN"

            db.sb.table("bodega_vendedores").insert({
                "bodega_id": bodega_id,
                "vendedor_id": vendedor_id,
                "rol": rol,
                "grupo": body.grupo,
                "supervisor": body.supervisor,
                "dia_visita": body.dia_visita,
                "dia_entrega": body.dia_entrega,
                "activo": True,
            }).execute()
            logger.info(f"Mapping creado: {bodega_id} -> {vendedor_id}")

    return {
        "ok": True,
        "bodega_id": bodega_id,
        "bodega_existe": bool(existing),
        "vendedor_id": vendedor_id,
        "mensaje": f"{'Actualizada' if existing else 'Creada'}: {body.razon_social} con línea S/{body.linea_aprobada}",
    }
