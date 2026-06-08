"""Importación masiva desde Excel (.xlsx) para backoffice Circa."""
from __future__ import annotations

import io
import json
import re
from typing import Any

from fastapi import HTTPException
from openpyxl import Workbook, load_workbook

from app.routes import distribuidor as dist
from app.services import db
from app.services.backoffice_audit import log_action
from app.services.distribuidor_routing import DIMAX_DISTRIBUIDOR_ID, ZOOM_DISTRIBUIDOR_ID

# ── Plantillas: encabezados ───────────────────────────────────────

BODEGAS_HEADERS = [
    "ruc",
    "razon_social",
    "nombre_comercial",
    "telefono_whatsapp",
    "representante_legal",
    "dni_representante",
    "linea_aprobada",
    "estado",
    "es_test",
    "solo_dni_sin_ruc",
    "direccion_fiscal",
    "distrito",
    "provincia",
]

BODEGAS_EJEMPLO = [
    [
        "20123456789",
        "BODEGA EJEMPLO SAC",
        "Minimarket El Sol",
        "999888777",
        "JUAN PEREZ GARCIA",
        "12345678",
        500,
        "preaprobada",
        0,
        0,
        "Av. Principal 123",
        "Miraflores",
        "Lima",
    ],
]

PEDIDOS_HEADERS = [
    "ref_pedido",
    "ruc_bodega",
    "tipo",
    "sku_distribuidor",
    "cantidad",
    "precio_unitario",
    "unidad",
    "descripcion",
    "descuento_prorrateado",
    "dimax_pedido_id",
    "vendedor_codigo",
    "vendedor_nombre",
    "telefono_whatsapp",
    "razon_social",
]

PEDIDOS_EJEMPLO = [
    ["P001", "20123456789", "preventa", "10001234", 2, 12.5, "UND x 1", "Producto demo", 0, "", "V001", "Carlos Vendedor", "", ""],
    ["P001", "20123456789", "preventa", "10005678", 1, 45.0, "CJA x 6", "Otro producto", 0, "", "V001", "Carlos Vendedor", "", ""],
]

HEADER_ALIASES: dict[str, tuple[str, ...]] = {
    "ruc": ("ruc", "ruc_bodega", "ruc bodega"),
    "razon_social": ("razon_social", "razon social", "razon"),
    "nombre_comercial": ("nombre_comercial", "nombre comercial", "comercial"),
    "telefono_whatsapp": ("telefono_whatsapp", "telefono", "whatsapp", "celular", "movil"),
    "representante_legal": ("representante_legal", "representante", "dueno"),
    "dni_representante": ("dni_representante", "dni", "dni_dueno"),
    "linea_aprobada": ("linea_aprobada", "linea", "linea aprobada"),
    "estado": ("estado",),
    "es_test": ("es_test", "test", "es prueba"),
    "solo_dni_sin_ruc": ("solo_dni_sin_ruc", "solo_dni", "sin_ruc"),
    "direccion_fiscal": ("direccion_fiscal", "direccion", "domicilio"),
    "distrito": ("distrito",),
    "provincia": ("provincia", "departamento"),
    "ref_pedido": ("ref_pedido", "ref pedido", "grupo", "pedido", "numero_pedido"),
    "ruc_bodega": ("ruc_bodega", "ruc", "ruc bodega"),
    "tipo": ("tipo", "tipo_operacion", "tipo operacion"),
    "sku_distribuidor": ("sku_distribuidor", "sku", "codigo", "codigo_sku"),
    "cantidad": ("cantidad", "qty", "cant"),
    "precio_unitario": ("precio_unitario", "precio", "precio unitario", "p_unit"),
    "unidad": ("unidad", "formato", "pack"),
    "descripcion": ("descripcion", "producto", "nombre"),
    "descuento_prorrateado": ("descuento_prorrateado", "descuento"),
    "dimax_pedido_id": ("dimax_pedido_id", "ref_externa", "id_dimax", "id_externo"),
    "vendedor_codigo": ("vendedor_codigo", "codigo_vendedor", "vendedor"),
    "vendedor_nombre": ("vendedor_nombre", "nombre_vendedor"),
}


def _norm_header(value: Any) -> str:
    s = str(value or "").strip().lower()
    s = re.sub(r"\s+", " ", s)
    return s


def _cell_str(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value == int(value):
        return str(int(value))
    return str(value).strip()


def _cell_float(value: Any, default: float = 0.0) -> float:
    if value is None or value == "":
        return default
    try:
        return float(value)
    except (TypeError, ValueError):
        return default


def _cell_bool(value: Any, default: bool = False) -> bool:
    if value is None or value == "":
        return default
    s = str(value).strip().lower()
    return s in ("1", "true", "si", "sí", "yes", "y", "x")


def _map_headers(raw_headers: list[Any]) -> dict[int, str]:
    """Mapea índice de columna → campo canónico."""
    canonical: dict[str, str] = {}
    for aliases in HEADER_ALIASES.values():
        canonical[aliases[0]] = aliases[0]
    for field, aliases in HEADER_ALIASES.items():
        for a in aliases:
            canonical[_norm_header(a)] = field

    mapping: dict[int, str] = {}
    for idx, h in enumerate(raw_headers):
        key = canonical.get(_norm_header(h))
        if key:
            mapping[idx] = key
    return mapping


def parse_xlsx(content: bytes) -> list[dict[str, Any]]:
    """Lee la primera hoja y devuelve filas como dicts (encabezado fila 1)."""
    try:
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Archivo Excel inválido: {e}") from e

    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration:
        raise HTTPException(status_code=400, detail="El Excel está vacío")

    col_map = _map_headers(list(header_row))
    if not col_map:
        raise HTTPException(status_code=400, detail="No se reconocieron columnas. Usa la plantilla oficial.")

    result: list[dict[str, Any]] = []
    for row_num, row in enumerate(rows_iter, start=2):
        if not row or all(c is None or str(c).strip() == "" for c in row):
            continue
        item: dict[str, Any] = {"_fila": row_num}
        for idx, field in col_map.items():
            if idx < len(row):
                item[field] = row[idx]
        result.append(item)
    wb.close()
    return result


def build_template_xlsx(headers: list[str], ejemplo: list[list[Any]]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Datos"
    ws.append(headers)
    for row in ejemplo:
        ws.append(row)
    # Hoja de ayuda
    help_ws = wb.create_sheet("Ayuda")
    help_ws.append(["Campo", "Descripción"])
    help_ws.append(["es_test", "0 = bodega real, 1 = bodega de prueba"])
    help_ws.append(["estado", "preaprobada | inactivo | activo"])
    help_ws.append(["ref_pedido", "Pedidos: mismo ref_pedido = un solo pedido con varias filas"])
    help_ws.append(["tipo", "preventa (default) | venta"])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def normalizar_telefono_peru(tel: str) -> str:
    t = "".join(c for c in str(tel) if c.isdigit() or c == "+")
    if t.startswith("+51"):
        return t
    if t.startswith("51") and len(t) == 11:
        return "+" + t
    if len(t) == 9:
        return "+51" + t
    if str(tel).strip().startswith("+"):
        return str(tel).strip()
    raise ValueError("Teléfono WhatsApp inválido (Perú)")


def _normalize_bodega_row(row: dict[str, Any], *, mode_test: bool | None = None) -> dict[str, Any]:
    ruc = _cell_str(row.get("ruc"))
    razon = _cell_str(row.get("razon_social"))
    tel_raw = _cell_str(row.get("telefono_whatsapp"))
    if not ruc or not razon or not tel_raw:
        raise ValueError("Faltan ruc, razon_social o telefono_whatsapp")

    tel = normalizar_telefono_peru(tel_raw)
    es_test = _cell_bool(row.get("es_test"), default=False)
    if mode_test is True:
        es_test = True
    elif mode_test is False:
        es_test = False

    linea = _cell_float(row.get("linea_aprobada"), 500.0)
    if linea <= 0:
        linea = 500.0

    return {
        "fila": row.get("_fila", "?"),
        "ruc": ruc,
        "razon_social": razon,
        "nombre_comercial": _cell_str(row.get("nombre_comercial")) or razon,
        "telefono_whatsapp": tel,
        "representante_legal": _cell_str(row.get("representante_legal")) or None,
        "dni_representante": _cell_str(row.get("dni_representante")) or None,
        "linea_aprobada": linea,
        "estado": _cell_str(row.get("estado")) or "preaprobada",
        "es_test": es_test,
        "solo_dni_sin_ruc": _cell_bool(row.get("solo_dni_sin_ruc")),
        "direccion_fiscal": _cell_str(row.get("direccion_fiscal")) or None,
        "distrito": _cell_str(row.get("distrito")) or None,
        "provincia": _cell_str(row.get("provincia")) or None,
    }


def preview_bodegas_rows(
    rows: list[dict[str, Any]],
    *,
    mode_test: bool | None = None,
) -> dict[str, Any]:
    """Valida filas sin insertar — para tabla de previsualización."""
    previews: list[dict[str, Any]] = []
    listas = omitidas = errores = 0

    for row in rows:
        fila = row.get("_fila", "?")
        status = "ok"
        issues: list[str] = []
        normalized: dict[str, Any] = {}

        try:
            normalized = _normalize_bodega_row(row, mode_test=mode_test)
            if db.get_bodega_by_ruc(normalized["ruc"]):
                status = "omitir"
                issues.append("RUC ya existe")
            elif db.get_bodega_by_phone(normalized["telefono_whatsapp"]):
                status = "error"
                issues.append("Teléfono ya registrado")
        except Exception as e:
            status = "error"
            issues.append(str(e))
            normalized = {
                "fila": fila,
                "ruc": _cell_str(row.get("ruc")),
                "razon_social": _cell_str(row.get("razon_social")),
                "nombre_comercial": _cell_str(row.get("nombre_comercial")),
                "telefono_whatsapp": _cell_str(row.get("telefono_whatsapp")),
                "linea_aprobada": _cell_float(row.get("linea_aprobada"), 500.0),
                "estado": _cell_str(row.get("estado")) or "preaprobada",
                "es_test": _cell_bool(row.get("es_test")),
            }

        if status == "ok":
            listas += 1
        elif status == "omitir":
            omitidas += 1
        else:
            errores += 1

        previews.append({
            **normalized,
            "status": status,
            "issues": issues,
            "can_import": status == "ok",
        })

    return {
        "rows": previews,
        "summary": {
            "total": len(previews),
            "listas": listas,
            "omitidas": omitidas,
            "errores": errores,
            "can_import": listas > 0,
        },
    }


def import_bodegas_rows(
    rows: list[dict[str, Any]],
    *,
    user: dict,
    comentario: str,
    mode_test: bool | None = None,
) -> dict[str, Any]:
    creadas = 0
    omitidas = 0
    errores: list[dict[str, Any]] = []

    for row in rows:
        fila = row.get("_fila", "?")
        try:
            normalized = _normalize_bodega_row(row, mode_test=mode_test)
            ruc = normalized["ruc"]

            if db.get_bodega_by_ruc(ruc):
                omitidas += 1
                errores.append({"fila": fila, "ruc": ruc, "error": "RUC ya existe (omitida)"})
                continue

            if db.get_bodega_by_phone(normalized["telefono_whatsapp"]):
                raise ValueError("Teléfono ya registrado en otra bodega")

            dist_id = ZOOM_DISTRIBUIDOR_ID if normalized["es_test"] else DIMAX_DISTRIBUIDOR_ID
            payload = {
                **normalized,
                "linea_disponible": 0,
                "distribuidor_id": dist_id,
                "en_piloto": True,
            }
            payload.pop("fila", None)
            ins = db.sb.table("bodegas").insert(payload).execute()
            bodega_id = ins.data[0]["id"] if ins.data else None
            creadas += 1
            log_action(
                user=user,
                action="bodega_import",
                entity_type="bodega",
                entity_id=bodega_id,
                comment=comentario,
                after={"ruc": ruc, "fila": fila},
                bodega_id=bodega_id,
            )
        except Exception as e:
            errores.append({"fila": fila, "ruc": _cell_str(row.get("ruc")), "error": str(e)})

    return {
        "ok": creadas > 0 and len(errores) == omitidas,
        "creadas": creadas,
        "omitidas": omitidas,
        "errores": errores,
        "total_filas": len(rows),
    }


def _group_pedido_rows(rows: list[dict[str, Any]]) -> dict[str, list[dict[str, Any]]]:
    groups: dict[str, list[dict[str, Any]]] = {}
    auto_counter = 0
    last_ruc = ""
    for row in rows:
        ref = _cell_str(row.get("ref_pedido"))
        ruc = _cell_str(row.get("ruc_bodega"))
        if not ref:
            if ruc != last_ruc:
                auto_counter += 1
                last_ruc = ruc
            ref = f"AUTO-{auto_counter}-{ruc}"
        groups.setdefault(ref, []).append(row)
    return groups


def _resolver_vendedor(codigo: str, nombre: str, distribuidor_id: str) -> str | None:
    if not codigo:
        return None
    return dist._resolver_o_crear_vendedor(codigo, nombre or codigo, distribuidor_id)


def import_pedidos_rows(
    rows: list[dict[str, Any]],
    *,
    user: dict,
    comentario: str,
    crear_bodega_si_falta: bool = False,
) -> dict[str, Any]:
    groups = _group_pedido_rows(rows)
    creados = 0
    errores: list[dict[str, Any]] = []
    pedidos_creados: list[dict[str, Any]] = []

    for ref, group_rows in groups.items():
        first = group_rows[0]
        fila = first.get("_fila", "?")
        try:
            ruc = _cell_str(first.get("ruc_bodega"))
            if not ruc:
                raise ValueError("Falta ruc_bodega")

            bodega = db.get_bodega_by_ruc(ruc)
            if not bodega and crear_bodega_si_falta:
                tel_raw = _cell_str(first.get("telefono_whatsapp"))
                razon = _cell_str(first.get("razon_social")) or f"Bodega {ruc}"
                if not tel_raw:
                    raise ValueError("Bodega no existe y falta telefono_whatsapp para crearla")
                tel = normalizar_telefono_peru(tel_raw)
                dist_id = DIMAX_DISTRIBUIDOR_ID
                bodega, _ = db.upsert_bodega_para_preventa(
                    ruc,
                    dist_id,
                    razon_social=razon,
                    nombre_comercial=razon,
                    telefono_whatsapp=tel,
                )
            if not bodega:
                raise ValueError(f"Bodega con RUC {ruc} no encontrada")

            bodega_id = bodega["id"]
            dist_id = db.get_distribuidor_pedido_de_bodega(bodega_id) or (
                ZOOM_DISTRIBUIDOR_ID if bodega.get("es_test") else DIMAX_DISTRIBUIDOR_ID
            )

            tipo = _cell_str(first.get("tipo")).lower() or "preventa"
            if tipo not in ("preventa", "venta"):
                raise ValueError("tipo debe ser preventa o venta")

            items_dimax = []
            items_json = []
            total = 0.0
            for gr in group_rows:
                sku = _cell_str(gr.get("sku_distribuidor"))
                qty = int(_cell_float(gr.get("cantidad"), 0))
                precio = _cell_float(gr.get("precio_unitario"), 0)
                if not sku or qty <= 0:
                    continue
                subtotal = qty * precio
                total += subtotal
                unidad = _cell_str(gr.get("unidad")) or "UND x 1"
                desc = _cell_str(gr.get("descripcion"))
                items_dimax.append({
                    "sku_distribuidor": sku,
                    "cantidad": qty,
                    "unidad": unidad,
                    "precio_unitario": precio,
                    "subtotal": subtotal,
                    "descripcion": desc or None,
                })

            if not items_dimax:
                raise ValueError("Sin ítems válidos (sku + cantidad)")

            descuento = _cell_float(first.get("descuento_prorrateado"), 0)
            vendedor_id = _resolver_vendedor(
                _cell_str(first.get("vendedor_codigo")),
                _cell_str(first.get("vendedor_nombre")),
                dist_id,
            )
            dimax_id = _cell_str(first.get("dimax_pedido_id")) or None

            if tipo == "preventa":
                resultado = db.crear_pedido_preventa(
                    bodega_id=bodega_id,
                    distribuidor_id=dist_id,
                    items_dimax=items_dimax,
                    total_pedido=max(0, total - descuento),
                    descuento_prorrateado=descuento,
                    vendedor_id=vendedor_id,
                    dimax_pedido_id=dimax_id,
                )
                pedido_id = resultado["pedido_id"]
                db.sb.table("pedidos").update({
                    "origen": "backoffice_excel",
                    "tipo_operacion": "preventa",
                }).eq("id", pedido_id).execute()
                pedidos_creados.append({
                    "ref_pedido": ref,
                    "pedido_id": pedido_id,
                    "ruc": ruc,
                    "items_creados": resultado["items_creados"],
                    "items_no_match": resultado["items_no_match"],
                })
            else:
                # Venta en borrador — items_json para flujo catálogo/PIN
                sku_to_prod = {}
                for c in cat_rows:
                    sku_to_prod[c["sku_distribuidor"]] = c

                for it in items_dimax:
                    sku_norm = str(it["sku_distribuidor"]).lstrip("0") or "0"
                    cat = sku_to_prod.get(sku_norm) or sku_to_prod.get(it["sku_distribuidor"])
                    pc = (cat or {}).get("productos_circa") or {}
                    pack = 1
                    try:
                        pack = int(str(it["unidad"]).split("x")[-1].strip())
                    except Exception:
                        pass
                    items_json.append({
                        "catalogo_id": (cat or {}).get("producto_circa_id"),
                        "nombre": pc.get("nombre") or it.get("descripcion") or sku_norm,
                        "marca": pc.get("marca") or "",
                        "pack_size": pack,
                        "cantidad": it["cantidad"],
                        "precio": it["precio_unitario"],
                        "subtotal": it["subtotal"],
                    })

                import json
                pedido_payload = {
                    "bodega_id": bodega_id,
                    "distribuidor_id": dist_id,
                    "vendedor_id": vendedor_id,
                    "items_json": json.dumps(items_json),
                    "monto_productos": total,
                    "total_pedido": max(0, total - descuento),
                    "descuento_prorrateado": descuento,
                    "estado": "borrador",
                    "tipo_operacion": "venta",
                    "origen": "backoffice_excel",
                }
                ins = db.sb.table("pedidos").insert(pedido_payload).execute()
                pedido_id = ins.data[0]["id"] if ins.data else None
                pedidos_creados.append({"ref_pedido": ref, "pedido_id": pedido_id, "ruc": ruc, "tipo": "venta_borrador"})

            creados += 1
            log_action(
                user=user,
                action="pedido_import",
                entity_type="pedido",
                entity_id=pedidos_creados[-1].get("pedido_id"),
                comment=comentario,
                after={"ref_pedido": ref, "ruc": ruc, "tipo": tipo},
                bodega_id=bodega_id,
                pedido_id=pedidos_creados[-1].get("pedido_id"),
            )
        except Exception as e:
            errores.append({"ref_pedido": ref, "fila": fila, "ruc": _cell_str(first.get("ruc_bodega")), "error": str(e)})

    return {
        "ok": creados > 0,
        "creados": creados,
        "errores": errores,
        "pedidos": pedidos_creados,
        "total_grupos": len(groups),
    }
