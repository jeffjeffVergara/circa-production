"""Parser del Excel DIMAX de alta de cliente (hoja ``clientes``)."""
from __future__ import annotations

import io
import re
from datetime import date, datetime
from typing import Any

from fastapi import HTTPException
from openpyxl import load_workbook

from app.services.excel_import import normalizar_telefono_peru

SHEET_CLIENTES = "clientes"

COL_CODIGO = "Codigo"
COL_DOC = "DNI/RUC"
COL_RAZON = "RazonSocial"
COL_DIRECCION = "Direccion"
COL_TELEFONO = "TELEFONO"
COL_DISTRITO = "Distrito"
COL_VENDEDOR_COD = "COD VENDEDOR 1"
COL_VENDEDOR_NOM = "VENDEDOR 1"


def _norm_header(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip().lower())


def _cell_str(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value == int(value):
        return str(int(value))
    if isinstance(value, (datetime, date)):
        return value.isoformat()[:10]
    return str(value).strip()


def _find_clientes_sheet(wb) -> Any:
    for name in wb.sheetnames:
        if _norm_header(name) == SHEET_CLIENTES:
            return wb[name]
    # Fallback: primera hoja con columnas DIMAX reconocibles
    for name in wb.sheetnames:
        ws = wb[name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        headers = {_norm_header(h) for h in rows[0] if h is not None}
        if "dni/ruc" in headers and "razonsocial" in headers:
            return ws
    raise HTTPException(
        status_code=400,
        detail="No se encontró la hoja «clientes» ni columnas DIMAX (DNI/RUC, RazonSocial).",
    )


def _parse_documento(raw: str) -> tuple[str, bool]:
    """Returns (documento_limpio, solo_dni_sin_ruc)."""
    doc = re.sub(r"\D", "", raw or "")
    if not doc:
        raise ValueError("Falta DNI/RUC")
    if len(doc) == 8:
        return doc, True
    if len(doc) == 11:
        return doc, False
    raise ValueError(f"DNI/RUC inválido ({len(doc)} dígitos): se espera 8 (DNI) u 11 (RUC)")


def parse_dimax_bodega_excel(content: bytes) -> dict[str, Any]:
    """Lee la primera fila de datos de la hoja clientes DIMAX."""
    try:
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Archivo Excel inválido: {e}") from e

    try:
        ws = _find_clientes_sheet(wb)
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 2:
            raise HTTPException(status_code=400, detail="El Excel no tiene filas de cliente")

        headers = [str(h).strip() if h is not None else "" for h in rows[0]]
        idx = {h: i for i, h in enumerate(headers)}

        def col(row: tuple, name: str, default=None):
            i = idx.get(name)
            if i is None or i >= len(row):
                return default
            return row[i]

        data_row = None
        data_fila = None
        for row_num, row in enumerate(rows[1:], start=2):
            if not row or all(c is None or str(c).strip() == "" for c in row):
                continue
            doc = _cell_str(col(row, COL_DOC))
            if doc:
                data_row = row
                data_fila = row_num
                break

        if data_row is None:
            raise HTTPException(status_code=400, detail="Sin fila de cliente con DNI/RUC")

        doc_raw = _cell_str(col(data_row, COL_DOC))
        ruc, solo_dni = _parse_documento(doc_raw)
        razon = _cell_str(col(data_row, COL_RAZON))
        if not razon:
            raise ValueError("Falta RazonSocial")

        tel_raw = _cell_str(col(data_row, COL_TELEFONO))
        if not tel_raw:
            raise ValueError("Falta TELEFONO")

        tel = normalizar_telefono_peru(tel_raw)
        direccion = _cell_str(col(data_row, COL_DIRECCION))
        distrito = _cell_str(col(data_row, COL_DISTRITO))

        return {
            "formato": "dimax_clientes",
            "fila": data_fila,
            "codigo_dimax": _cell_str(col(data_row, COL_CODIGO)),
            "ruc": ruc,
            "solo_dni_sin_ruc": solo_dni,
            "razon_social": razon,
            "nombre_comercial": razon,
            "representante_legal": razon,
            "dni_representante": ruc if solo_dni else None,
            "telefono_whatsapp": tel,
            "direccion_fiscal": direccion or None,
            "distrito": distrito or None,
            "provincia": "Lima" if distrito else None,
            "vendedor_codigo": _cell_str(col(data_row, COL_VENDEDOR_COD)) or None,
            "vendedor_nombre": _cell_str(col(data_row, COL_VENDEDOR_NOM)) or None,
            "estado": "preaprobada",
            "linea_aprobada": 500.0,
            "linea_disponible": 0.0,
            "es_test": False,
        }
    finally:
        wb.close()


def enrich_preview(preview: dict[str, Any]) -> dict[str, Any]:
    """Añade alertas de duplicados consultando BD."""
    from app.services import db

    warnings: list[str] = []
    exists_ruc = False
    exists_phone = False

    ruc = preview.get("ruc") or ""
    if ruc and db.get_bodega_by_ruc(ruc):
        exists_ruc = True
        warnings.append(f"Ya existe una bodega con documento {ruc}")

    tel = preview.get("telefono_whatsapp") or ""
    if tel and db.get_bodega_by_phone(tel):
        exists_phone = True
        warnings.append(f"Ya existe una bodega con teléfono {tel}")

    out = dict(preview)
    out["warnings"] = warnings
    out["exists_ruc"] = exists_ruc
    out["exists_phone"] = exists_phone
    out["can_create"] = not exists_ruc and not exists_phone
    return out
