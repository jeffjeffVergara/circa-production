"""Lectura del Excel DIMAX (cliente + historial)."""

from __future__ import annotations

import io
import os
from typing import Any

import openpyxl


def _leer_workbook(wb, archivo: str) -> dict[str, Any]:
    hojas_cli = [s for s in wb.sheetnames if "cliente" in s.lower()]
    ws = wb[hojas_cli[0]] if hojas_cli else wb.worksheets[0]
    filas = list(ws.iter_rows(values_only=True))
    if len(filas) < 2:
        raise ValueError("La hoja de cliente no tiene datos")
    cli = dict(zip(filas[0], filas[1]))
    hojas_h = [s for s in wb.sheetnames if "istorial" in s.lower()]
    if not hojas_h:
        raise ValueError("No se encontro la hoja de Historial")
    hist = list(wb[hojas_h[0]].iter_rows(values_only=True))[1:]
    return {"cliente": cli, "historial": hist, "archivo": archivo}


def leer_bodega(path: str) -> dict[str, Any]:
    wb = openpyxl.load_workbook(path, data_only=True)
    return _leer_workbook(wb, os.path.basename(path))


def leer_bodega_bytes(content: bytes, filename: str) -> dict[str, Any]:
    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    return _leer_workbook(wb, filename or "upload.xlsx")
