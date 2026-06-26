#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
cargar_bodega.py  v2 - Carga de bodegas para el piloto Circa / DIMAX

QUE HACE
  Lee el/los Excel de historial de bodega(s) de DIMAX y, por cada una:
    - analiza el comportamiento de compra de los ultimos 6 meses
    - calcula la linea de credito con el modelo de riesgo (tier que cubre 7d)
    - en casos limite SIEMPRE asigna el tier conservador (no sube)
    - alerta si el ciclo de compra es largo o irregular (>10 dias o errático)
    - genera el SQL de creacion y el mensaje de WhatsApp de enrolamiento
    - marca solo lo que de verdad necesita criterio humano

DOS MODOS
  1) GENERAR (por defecto, no toca la base):
       python3 cargar_bodega.py carpeta/
     Crea salida_carga/reporte_bodegas.md y salida_carga/carga_bodegas.sql

  2) EJECUTAR (carga en la base, con confirmacion):
       python3 cargar_bodega.py carpeta/ --ejecutar
     Pide confirmacion antes de tocar la base. Las bodegas limpias se
     confirman en lote; las marcadas para revisar, una por una.
     Necesita la variable de entorno CIRCA_DB_URL (ver instrucciones abajo).
"""

import sys
import os
import glob
import statistics
from datetime import timedelta
from collections import defaultdict

try:
    import openpyxl
except ImportError:
    print("Falta openpyxl. Instalalo con:  pip3 install openpyxl")
    sys.exit(1)

# =====================================================================
# CONSTANTES DEL PILOTO
# =====================================================================
DIMAX_ID = "d1a2b3c4-0001-4000-8000-000000000002"
TIERS = [100, 200, 300, 400, 500]
VENTANA_DIAS = 182
LINK_ONBOARDING = "https://wa.me/51986311567?text=Hola"
MIN_PEDIDOS_HISTORIAL = 6
CICLO_LARGO_DIAS = 10        # compra cada mas dias que esto -> alerta de ciclo largo
CV_IRREGULAR = 0.70          # coef. de variacion mayor a esto -> alerta de irregular
DB_ENV_VAR = "CIRCA_DB_URL"

PREPOSICIONES = {"de", "del", "la", "las", "los", "y", "en"}


# =====================================================================
# HELPERS
# =====================================================================
def tier_para(linea_7d):
    """Tier conservador: el minimo (100..500) que cubre la linea de 7 dias."""
    for t in TIERS:
        if linea_7d <= t:
            return t
    return 500


def titulo(texto):
    if not texto:
        return ""
    out = []
    for i, p in enumerate(str(texto).strip().split()):
        bajo = p.lower()
        out.append(bajo if (i > 0 and bajo in PREPOSICIONES) else p.capitalize())
    return " ".join(out)


def doc_info(docnum):
    d = str(docnum).strip().replace(".0", "")
    if len(d) == 11 and d.startswith("10"):
        return {"ruc": d, "dni": d[2:10], "solo_dni": False}
    if len(d) == 11:
        return {"ruc": d, "dni": None, "solo_dni": False}
    # DNI puro. La columna dni_representante es VARCHAR(8) en la base.
    # Si el Excel trae el DNI con ceros de padding al frente (>8 chars),
    # los quitamos. Si trae menos de 8, rellenamos con zfill.
    dni = d.lstrip("0") if len(d) > 8 else d
    dni = dni.zfill(8)
    if len(dni) != 8:
        raise ValueError(
            "DNI invalido, no se puede normalizar a 8 caracteres: "
            "'%s' -> '%s'. Revisa el dato en el Excel." % (docnum, dni)
        )
    return {"ruc": None, "dni": dni, "solo_dni": True}


def telefono_e164(tel):
    t = str(tel).strip().replace(" ", "").replace("-", "").replace(".0", "")
    if t.startswith("+51"):
        return t
    if t.startswith("51") and len(t) == 11:
        return "+" + t
    return "+51" + t


def primer_nombre(razon_social):
    p = str(razon_social).strip().split()
    if len(p) >= 3:
        return p[2].capitalize()
    return p[-1].capitalize() if p else ""


def normaliza_grupo(g):
    if not g:
        return None
    g = str(g).strip().upper()
    if g.startswith("CONF") and not g.startswith("CONFITERIA"):
        g = g.replace("CONF", "CONFITERIA", 1)
    return "GV - " + g


def rol_de(grupo):
    return "CONFITERIA" if "CONF" in str(grupo or "").upper() else "ABN"


def dia_min(d):
    return str(d or "").strip().lower()


def etiqueta_regularidad(a):
    """Etiqueta legible de regularidad de compra a partir del CV."""
    cv = a.get("cv")
    if cv is None:
        return "sin datos suficientes"
    estado = "irregular" if cv > CV_IRREGULAR else "regular"
    return "%s (CV %.2f)" % (estado, cv)


def sql_str(valor):
    if valor is None:
        return "NULL"
    return "'" + str(valor).replace("'", "''") + "'"


# =====================================================================
# LECTURA DEL EXCEL
# =====================================================================
def leer_bodega(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    hojas_cli = [s for s in wb.sheetnames if "cliente" in s.lower()]
    ws = wb[hojas_cli[0]] if hojas_cli else wb.worksheets[0]
    filas = list(ws.iter_rows(values_only=True))
    if len(filas) < 2:
        raise ValueError("La hoja de cliente no tiene datos")
    cli = dict(zip(filas[0], filas[1]))
    hojas_h = [s for s in wb.sheetnames if "istorial" in s.lower()]
    if not hojas_h:
        raise ValueError("No se encontro la hoja de Historial")
    hist = list(wb[hojas_h[0]].iter_rows(values_only=True))[1:]
    return {"cliente": cli, "historial": hist, "archivo": os.path.basename(path)}


# =====================================================================
# ANALISIS DE RIESGO
# =====================================================================
def analizar(hist):
    por_fecha = defaultdict(float)
    for r in hist:
        if len(r) < 5:
            continue
        fecha, sellout = r[2], r[4]
        if fecha is not None and sellout is not None:
            por_fecha[fecha] += float(sellout)
    fechas = sorted(por_fecha)
    if not fechas:
        return None
    ultima = fechas[-1]
    corte = ultima - timedelta(days=VENTANA_DIAS)
    p6 = {f: v for f, v in por_fecha.items() if f >= corte}
    f6 = sorted(p6)
    n6 = len(f6)
    if n6 == 0:
        return None
    total6 = sum(p6.values())
    ticket = total6 / n6
    cv = None
    if n6 > 1:
        difs = [(f6[i + 1] - f6[i]).days for i in range(len(f6) - 1)]
        dias_entre = sum(difs) / len(difs)
        if len(difs) >= 2 and dias_entre:
            cv = statistics.pstdev(difs) / dias_entre
    else:
        dias_entre = float(VENTANA_DIAS)
    consumo_diario = ticket / dias_entre if dias_entre else ticket
    linea_7d = consumo_diario * 7
    return {
        "desde": corte, "hasta": ultima, "pedidos": n6, "total": total6,
        "ticket": ticket, "dias_entre": dias_entre, "cv": cv,
        "consumo_diario": consumo_diario, "linea_7d": linea_7d,
        "tier": tier_para(linea_7d),
    }


def clasificar_avisos(a):
    """Separa lo que NECESITA decision humana (revisar) de lo informativo (notas).
    Regla del piloto: en casos limite el tier es SIEMPRE el conservador,
    asi que 'ticket supera linea' es nota, no decision."""
    revisar, notas = [], []

    if a["pedidos"] < MIN_PEDIDOS_HISTORIAL:
        revisar.append(
            "Historial corto: solo %d pedidos en 6 meses. Decidir si hay "
            "comportamiento suficiente para asignar linea." % a["pedidos"])

    if a["linea_7d"] > 500:
        revisar.append(
            "El consumo de 7 dias (S/%.2f) supera el tier maximo de S/500. "
            "Bodega de alto volumen: revisar manualmente." % a["linea_7d"])

    if a["dias_entre"] > CICLO_LARGO_DIAS:
        revisar.append(
            "Ciclo de compra largo: compra cada ~%.0f dias, bastante mas que "
            "el plazo de credito de 7 dias. Bodega poco frecuente: monitorear "
            "de cerca el comportamiento de pago." % a["dias_entre"])

    if a.get("cv") is not None and a["cv"] > CV_IRREGULAR:
        revisar.append(
            "Compra irregular: los intervalos entre pedidos varian mucho "
            "(coef. de variacion %.2f; lo regular es por debajo de %.2f). "
            "Comportamiento poco predecible: monitorear de cerca."
            % (a["cv"], CV_IRREGULAR))

    if a["ticket"] > a["tier"]:
        notas.append(
            "Ticket promedio (S/%.2f) mayor que la linea (S/%d). Por politica "
            "conservadora se mantiene el tier; la bodega pagara la diferencia "
            "en efectivo. Sube de tier cuando tenga historial de pago."
            % (a["ticket"], a["tier"]))

    return {"revisar": revisar, "notas": notas}


# =====================================================================
# GENERACION DE SQL  (devuelve componentes separados)
# =====================================================================
def generar_sql(b):
    c = b["cliente"]
    a = b["analisis"]
    doc = doc_info(c.get("DNI/RUC"))
    tel = telefono_e164(c.get("TELEFONO"))
    razon = str(c.get("RazonSocial", "")).strip()
    direccion = titulo(c.get("Direccion"))
    distrito = titulo(c.get("Distrito"))
    tier = a["tier"]

    vendedores = []
    for n in ("", " 1", " 2"):
        cod = c.get("COD VENDEDOR" + n)
        if not cod:
            continue
        grupo = c.get("GRUPO" + n)
        vendedores.append({
            "codigo": str(cod).strip(),
            "nombre": str(c.get("VENDEDOR" + n, "")).strip(),
            "rol": rol_de(grupo), "grupo": normaliza_grupo(grupo),
            "supervisor": str(c.get("SUPERVISOR" + n, "")).strip(),
            "dia_visita": dia_min(c.get("DIA VISITA" + n)),
            "dia_entrega": dia_min(c.get("DIA ENTREGA" + n)),
        })

    ins = []
    for v in vendedores:
        ins.append("-- Vendedor %s (%s) - se crea solo si no existe"
                    % (v["codigo"], v["nombre"]))
        ins.append("INSERT INTO vendedores (distribuidor_id, codigo, nombre, activo)")
        ins.append("SELECT %s, %s, %s, true" % (
            sql_str(DIMAX_ID), sql_str(v["codigo"]), sql_str(v["nombre"])))
        ins.append("WHERE NOT EXISTS (SELECT 1 FROM vendedores")
        ins.append("  WHERE codigo = %s AND distribuidor_id = %s);"
                    % (sql_str(v["codigo"]), sql_str(DIMAX_ID)))
        ins.append("")

    ins.append("-- Crear la bodega (estado inactivo, disponible 0 hasta onboarding)")
    ins.append("INSERT INTO bodegas (")
    ins.append("  distribuidor_id, razon_social, nombre_comercial, telefono_whatsapp,")
    ins.append("  ruc, dni_representante, solo_dni_sin_ruc,")
    ins.append("  direccion_fiscal, direccion_despacho, distrito,")
    ins.append("  es_test, en_piloto, estado, linea_aprobada, linea_disponible)")
    ins.append("SELECT %s, %s, %s, %s," % (
        sql_str(DIMAX_ID), sql_str(razon), sql_str(razon), sql_str(tel)))
    ins.append("       %s, %s, %s," % (
        sql_str(doc["ruc"]), sql_str(doc["dni"]),
        "true" if doc["solo_dni"] else "false"))
    ins.append("       %s, %s, %s," % (
        sql_str(direccion), sql_str(direccion), sql_str(distrito)))
    ins.append("       false, true, 'inactivo', %d, 0" % tier)
    ins.append("WHERE NOT EXISTS (")
    ins.append("  SELECT 1 FROM bodegas WHERE telefono_whatsapp = %s);" % sql_str(tel))
    ins.append("")

    if vendedores:
        ins.append("-- Mapear vendedores a la bodega")
        ins.append("INSERT INTO bodega_vendedores")
        ins.append("  (bodega_id, vendedor_id, rol, grupo, supervisor,"
                    " dia_visita, dia_entrega, activo)")
        ins.append("SELECT b.id, v.id, t.rol, t.grupo, t.supervisor,"
                    " t.dia_visita, t.dia_entrega, true")
        ins.append("FROM (VALUES")
        filas_v = ["  (%s, %s, %s, %s, %s, %s)" % (
            sql_str(v["codigo"]), sql_str(v["rol"]), sql_str(v["grupo"]),
            sql_str(v["supervisor"]), sql_str(v["dia_visita"]),
            sql_str(v["dia_entrega"])) for v in vendedores]
        ins.append(",\n".join(filas_v))
        ins.append(") AS t(vendedor_codigo, rol, grupo, supervisor,"
                    " dia_visita, dia_entrega)")
        ins.append("JOIN bodegas b ON b.telefono_whatsapp = %s" % sql_str(tel))
        ins.append("JOIN vendedores v ON v.codigo = t.vendedor_codigo")
        ins.append("              AND v.distribuidor_id = %s" % sql_str(DIMAX_ID))
        ins.append("              AND v.activo = true")
        ins.append("WHERE NOT EXISTS (SELECT 1 FROM bodega_vendedores bv")
        ins.append("  WHERE bv.bodega_id = b.id AND bv.vendedor_id = v.id);")

    verif = "\n".join([
        "SELECT 'bodega' AS tipo, razon_social AS detalle,",
        "       linea_aprobada::text AS aprob, linea_disponible::text AS disp,",
        "       estado::text AS estado",
        "FROM bodegas WHERE telefono_whatsapp = %s" % sql_str(tel),
        "UNION ALL",
        "SELECT 'mapping', b.razon_social || ' -> ' || v.codigo,",
        "       bv.rol, bv.grupo, bv.dia_visita",
        "FROM bodega_vendedores bv",
        "JOIN bodegas b ON b.id = bv.bodega_id",
        "JOIN vendedores v ON v.id = bv.vendedor_id",
        "WHERE b.telefono_whatsapp = %s;" % sql_str(tel),
    ])

    return {"inserts": "\n".join(ins), "verificacion": verif,
            "vendedores": vendedores, "telefono": tel}


def sql_para_archivo(b):
    """Bloque SQL completo con BEGIN/COMMIT, para correr a mano en Supabase."""
    s = b["sql"]
    cab = ("-- ====================================================\n"
           "-- Bodega: %s\n"
           "-- Linea aprobada: S/%d  (modelo: consumo 7d = S/%.2f)\n"
           "-- ====================================================\n"
           % (str(b["cliente"].get("RazonSocial", "")).strip(),
              b["analisis"]["tier"], b["analisis"]["linea_7d"]))
    return cab + "BEGIN;\n\n" + s["inserts"] + "\n\n" + s["verificacion"] + "\n\nCOMMIT;\n"


# =====================================================================
# MENSAJE DE WHATSAPP
# =====================================================================
def generar_mensaje(b):
    c = b["cliente"]
    don = primer_nombre(c.get("RazonSocial", ""))
    vends = b["sql"]["vendedores"]
    v1 = vends[0] if vends else None
    nombre_vend = primer_nombre(v1["nombre"]) if v1 else "tu vendedor"
    dia = v1["dia_visita"] if v1 else "la semana"
    return (
        "Buenas Don %s! \U0001F44B Le escribe %s, de DIMAX.\n\n"
        "Le tengo una novedad para su bodega: ahora puede hacer sus pedidos "
        "por WhatsApp con Circa. Ve el catalogo completo, arma su pedido "
        "cuando quiera y lo recibe igual que siempre - sin tener que esperar "
        "a mi visita del %s.\n\n"
        "Y por su buen historial como cliente, ya le tenemos una linea de "
        "credito pre-aprobada \U0001F64C Para que pueda surtir su bodega y "
        "pagar con calma.\n\n"
        "Activarla le toma 2 minutos. Solo abra este enlace y envie el "
        "mensaje que le aparece:\n\U0001F449 %s\n\n"
        "Cualquier duda me avisa. Saludos!"
        % (don, nombre_vend, dia, LINK_ONBOARDING))


# =====================================================================
# FICHA DEL REPORTE
# =====================================================================
def generar_ficha(b):
    c, a = b["cliente"], b["analisis"]
    L = ["## %s" % str(c.get("RazonSocial", "")).strip(), ""]
    L.append("- Archivo: `%s`" % b["archivo"])
    L.append("- Codigo: %s  |  Documento: %s  |  Clasificacion: %s"
             % (c.get("Codigo"), c.get("DNI/RUC"), c.get("Clasificacion")))
    L.append("")
    L.append("### Analisis de riesgo (ultimos 6 meses)")
    L.append("")
    L.append("| Metrica | Valor |")
    L.append("|---|---|")
    L.append("| Periodo | %s -> %s |" % (a["desde"].date(), a["hasta"].date()))
    L.append("| Pedidos | %d |" % a["pedidos"])
    L.append("| Total comprado | S/%.2f |" % a["total"])
    L.append("| Ticket promedio | S/%.2f |" % a["ticket"])
    L.append("| Dias entre pedidos | %.1f |" % a["dias_entre"])
    L.append("| Regularidad de compra | %s |" % etiqueta_regularidad(a))
    L.append("| Consumo diario | S/%.2f |" % a["consumo_diario"])
    L.append("| Linea necesaria 7 dias | S/%.2f |" % a["linea_7d"])
    L.append("| **Tier asignado (conservador)** | **S/%d** |" % a["tier"])
    L.append("")
    if b["avisos"]["revisar"]:
        L.append("### ⚠ Necesita tu decision")
        L.append("")
        for x in b["avisos"]["revisar"]:
            L.append("- %s" % x)
        L.append("")
    if b["avisos"]["notas"]:
        L.append("### Notas")
        L.append("")
        for x in b["avisos"]["notas"]:
            L.append("- %s" % x)
        L.append("")
    L.append("### SQL de carga")
    L.append("")
    L.append("```sql")
    L.append(sql_para_archivo(b))
    L.append("```")
    L.append("")
    L.append("### Mensaje de WhatsApp")
    L.append("")
    L.append("```")
    L.append(b["mensaje"])
    L.append("```")
    L.append("")
    L.append("---")
    L.append("")
    return "\n".join(L)


# =====================================================================
# EJECUCION EN LA BASE  (modo --ejecutar)
# =====================================================================
def ejecutar_bodega(conn, b):
    """Corre los INSERT de una bodega y devuelve (ok, resultado_verificacion)."""
    cur = conn.cursor()
    try:
        cur.execute(b["sql"]["inserts"])
        conn.commit()
        cur.execute(b["sql"]["verificacion"])
        filas = cur.fetchall()
        cur.close()
        return True, filas
    except Exception as e:
        conn.rollback()
        cur.close()
        return False, str(e)


def correr_modo_ejecutar(procesadas):
    """Carga en la base con doble confirmacion."""
    db_url = os.environ.get(DB_ENV_VAR)
    if not db_url:
        print("\nNo se encontro la variable de entorno %s." % DB_ENV_VAR)
        print("El modo --ejecutar necesita la cadena de conexion de Supabase.")
        print("Configurala antes de correr (NO la pongas dentro del script):")
        print('    export %s="postgresql://usuario:clave@host:5432/postgres"'
              % DB_ENV_VAR)
        sys.exit(1)
    try:
        import psycopg2  # noqa
    except ImportError:
        print("\nEl modo --ejecutar necesita psycopg2. Instalalo con:")
        print("    pip3 install psycopg2-binary")
        sys.exit(1)
    import psycopg2

    limpias = [b for b in procesadas if not b["avisos"]["revisar"]]
    a_revisar = [b for b in procesadas if b["avisos"]["revisar"]]
    aprobadas = []

    if limpias:
        print("\nBodegas listas para cargar (sin observaciones):")
        for b in limpias:
            print("  - %-40s linea S/%d"
                  % (str(b["cliente"].get("RazonSocial", ""))[:40],
                     b["analisis"]["tier"]))
        r = input("\nEscribi CARGAR para cargar estas %d bodegas (o ENTER "
                  "para saltarlas): " % len(limpias)).strip()
        if r == "CARGAR":
            aprobadas += limpias
        else:
            print("  -> se omiten las bodegas limpias.")

    for b in a_revisar:
        nombre = str(b["cliente"].get("RazonSocial", "")).strip()
        print("\n--- REVISAR: %s (linea S/%d) ---"
              % (nombre, b["analisis"]["tier"]))
        for x in b["avisos"]["revisar"]:
            print("  ! %s" % x)
        r = input("Cargar esta bodega igual? (si / no): ").strip().lower()
        if r == "si":
            aprobadas += [b]
        else:
            print("  -> omitida.")

    if not aprobadas:
        print("\nNo se aprobo ninguna bodega. Nada que cargar.")
        return

    print("\nConectando a la base...")
    conn = psycopg2.connect(db_url)
    ok, fallo = 0, 0
    for b in aprobadas:
        nombre = str(b["cliente"].get("RazonSocial", "")).strip()
        exito, res = ejecutar_bodega(conn, b)
        if exito:
            ok += 1
            print("\n  OK  %s" % nombre)
            for fila in res:
                print("       " + " | ".join(str(x) for x in fila))
        else:
            fallo += 1
            print("\n  ERROR  %s" % nombre)
            print("       %s" % res)
    conn.close()
    print("\n=== Carga terminada: %d cargada(s), %d con error ==="
          % (ok, fallo))


# =====================================================================
# MAIN
# =====================================================================
def juntar_archivos(args):
    archivos = []
    for a in args:
        if os.path.isdir(a):
            archivos += sorted(glob.glob(os.path.join(a, "*.xlsx")))
        elif a.lower().endswith(".xlsx"):
            archivos.append(a)
    return archivos


def main():
    args = [a for a in sys.argv[1:] if not a.startswith("--")]
    modo_ejecutar = "--ejecutar" in sys.argv[1:]

    if not args:
        print(__doc__)
        sys.exit(0)

    archivos = juntar_archivos(args)
    if not archivos:
        print("No se encontraron archivos .xlsx en lo indicado.")
        sys.exit(1)

    procesadas, errores, vistos = [], [], {}

    for path in archivos:
        nombre = os.path.basename(path)
        try:
            b = leer_bodega(path)
        except Exception as e:
            errores.append((nombre, str(e)))
            continue
        codigo = str(b["cliente"].get("Codigo", "")).strip()
        if codigo and codigo in vistos:
            errores.append((nombre, "DUPLICADO: codigo %s ya estaba en %s. "
                            "Se omite." % (codigo, vistos[codigo])))
            continue
        if codigo:
            vistos[codigo] = nombre
        a = analizar(b["historial"])
        if not a:
            errores.append((nombre, "Sin historial de compras utilizable."))
            continue
        b["analisis"] = a
        b["avisos"] = clasificar_avisos(a)
        b["sql"] = generar_sql(b)
        b["mensaje"] = generar_mensaje(b)
        procesadas.append(b)

    # archivos de salida (siempre se generan)
    os.makedirs("salida_carga", exist_ok=True)
    rep = ["# Circa - Reporte de carga de bodegas", "",
           "Generado por cargar_bodega.py  |  %d bodega(s) procesada(s)"
           % len(procesadas), "", "---", ""]
    for b in procesadas:
        rep.append(generar_ficha(b))
    with open("salida_carga/reporte_bodegas.md", "w", encoding="utf-8") as f:
        f.write("\n".join(rep))
    with open("salida_carga/carga_bodegas.sql", "w", encoding="utf-8") as f:
        f.write("-- Circa - SQL de carga de bodegas\n")
        f.write("-- Cada bodega es un bloque BEGIN/COMMIT independiente.\n\n")
        for b in procesadas:
            f.write(sql_para_archivo(b) + "\n\n")

    # resumen en consola
    print("=" * 60)
    print("PROCESADAS: %d bodega(s)" % len(procesadas))
    print("=" * 60)
    for b in procesadas:
        marca = "  [REVISAR]" if b["avisos"]["revisar"] else ""
        print("  %-40s -> linea S/%d%s"
              % (str(b["cliente"].get("RazonSocial", ""))[:40],
                 b["analisis"]["tier"], marca))
    if errores:
        print("\nAVISOS / OMITIDOS:")
        for nombre, msg in errores:
            print("  %s: %s" % (nombre, msg))
    print("\nSalida en  salida_carga/  (reporte_bodegas.md  y  carga_bodegas.sql)")

    if modo_ejecutar:
        if procesadas:
            correr_modo_ejecutar(procesadas)
    else:
        print("\nPara cargar en la base con confirmacion:  agrega  --ejecutar")


if __name__ == "__main__":
    main()
